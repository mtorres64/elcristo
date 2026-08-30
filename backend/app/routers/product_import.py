"""Importación masiva de productos desde una planilla .xlsx.

- `GET  /products/import/template`  → descarga la plantilla vacía.
- `POST /products/import`           → sube la planilla y arranca un job en segundo plano.
- `GET  /products/import/{job_id}`  → estado/progreso del job (para el polling del modal).

Todas las filas se importan como la **medida mediana**: se llenan los campos base del
producto. No se crean variantes de tamaño (pequeña/grande), macetas recomendadas ni
imágenes. Si un producto ya existe (mismo SKU, o mismo título) se actualiza en lugar
de duplicarse.
"""

import io
from datetime import UTC, datetime

from bson import ObjectId
from fastapi import APIRouter, BackgroundTasks, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.schemas.product_import import ImportJobOut, ImportStartResponse
from app.utils.auth_deps import require_user_id
from app.utils.slugify import slugify

router = APIRouter()

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Encabezado de la plantilla (snake_case) → clave interna que usa el parser.
COLUMN_MAP: dict[str, str] = {
    "nombre": "title",
    "descripcion_corta": "short_description",
    "descripcion": "description",
    "categoria": "category",
    "precio": "price",
    "precio_promocional": "compare_at_price",
    "precio_costo": "cost_price",
    "moneda": "currency",
    "impuesto": "tax",
    "stock": "stock",
    "sku": "sku",
    "estado": "status",
    "destacado": "is_featured",
    "peso_kg": "weight_kg",
    "altura_cm": "height_cm",
    "etiquetas": "tags",
    "cuidado_luz": "care_light",
    "cuidado_riego": "care_water",
    "cuidado_ambiente": "care_environment",
    "cuidado_temperatura": "care_temperature",
    "diametro_maceta": "attr_pot_diameter",
    "altura_con_maceta": "attr_height_with_pot",
    "tipo_planta": "attr_plant_type",
    "crecimiento": "attr_growth",
}

TEMPLATE_HEADERS = list(COLUMN_MAP.keys())

# Columnas que en el panel son un <select> con opciones cortas y sin comas:
# se ofrecen como lista desplegable en la celda con la lista embebida.
DROPDOWN_CHOICES: dict[str, list[str]] = {
    "moneda": ["ARS", "USD"],
    "impuesto": ["iva-21", "iva-10", "exento"],
    "estado": ["activo", "borrador"],
    "destacado": ["si", "no"],
    "crecimiento": ["Lento", "Medio", "Rápido"],
}

# Campos de "cuidados": en el panel son chips sugeridos + texto libre. En la
# plantilla se ofrecen como desplegable (sin bloquear texto libre). Las opciones
# tienen comas, así que van por una hoja auxiliar. Deben coincidir con
# CARE_OPTIONS de frontend/src/pages/admin/ProductEdit.tsx.
CARE_CHOICES: dict[str, list[str]] = {
    "cuidado_luz": [
        "Luz directa intensa",
        "Luz brillante indirecta",
        "Luz brillante indirecta, tolera sombra parcial",
        "Luz indirecta moderada",
        "Sombra parcial",
        "Sombra total",
    ],
    "cuidado_riego": [
        "Abundante (casi diario)",
        "Frecuente en verano, reducir en invierno",
        "Moderado, dejar secar entre riegos",
        "Escaso, tolerante a la sequía",
        "Mínimo (cactus y suculentas)",
    ],
    "cuidado_ambiente": [
        "Ambientes con buena ventilación",
        "Interior luminoso",
        "Interior con poca luz",
        "Exterior protegido",
        "Interior o exterior",
        "Alta humedad ambiental",
    ],
    "cuidado_temperatura": [
        "5° - 15°C (resistente al frío)",
        "10° - 25°C (templado)",
        "15° - 35°C / Evitar heladas",
        "20° - 40°C (tropical)",
        "Resistente a heladas",
    ],
}

# Hasta qué fila se aplican los desplegables de la plantilla.
TEMPLATE_VALIDATION_ROWS = 500

TEMPLATE_EXAMPLE = {
    "nombre": "Pothos Dorado",
    "descripcion_corta": "La planta más resistente del mundo. Ideal para principiantes.",
    "descripcion": "Trepadora de interior muy tolerante a la falta de luz y riego irregular.",
    "categoria": "Plantas de interior",
    "precio": 6800,
    "precio_promocional": 8900,
    "precio_costo": 3200,
    "moneda": "ARS",
    "impuesto": "iva-21",
    "stock": 60,
    "sku": "PLT-POT-DOR-001",
    "estado": "activo",
    "destacado": "no",
    "peso_kg": 1.2,
    "altura_cm": 30,
    "etiquetas": "interior, principiantes, colgante",
    "cuidado_luz": "Luz media a sombra parcial",
    "cuidado_riego": "Moderado, dejar secar la capa superior",
    "cuidado_ambiente": "Interiores cálidos, tolera baja humedad",
    "cuidado_temperatura": "15° - 30°C",
    "diametro_maceta": "17 cm",
    "altura_con_maceta": "25 - 35 cm",
    "tipo_planta": "Trepadora de interior",
    "crecimiento": "Rápido",
}

INSTRUCTIONS = [
    ("nombre", "Obligatorio. Nombre del producto."),
    ("descripcion_corta", "Opcional. Máximo 160 caracteres."),
    ("descripcion", "Opcional. Descripción completa."),
    ("categoria", "Opcional. Nombre de una categoría existente (ver lista abajo). "
                  "Si no coincide, el producto se importa sin categoría."),
    ("precio", "Obligatorio. Precio de venta en pesos (ej. 6800), medida mediana."),
    ("precio_promocional", "Opcional. Precio tachado / de comparación, en pesos."),
    ("precio_costo", "Opcional. Precio de costo en pesos."),
    ("moneda", "ARS (por defecto) o USD."),
    ("impuesto", "iva-21 (por defecto), iva-10 o exento."),
    ("stock", "Cantidad disponible (entero). Por defecto 0."),
    ("sku", "Opcional. Se usa para detectar productos ya existentes y actualizarlos."),
    ("estado", "activo o borrador. Vacío = borrador."),
    ("destacado", "si o no. Por defecto no."),
    ("peso_kg", "Opcional. Peso en kilogramos (ej. 1.2)."),
    ("altura_cm", "Opcional. Altura de la planta en centímetros (entero)."),
    ("etiquetas", "Opcional. Varias separadas por coma."),
    ("cuidado_luz", "Opcional."),
    ("cuidado_riego", "Opcional."),
    ("cuidado_ambiente", "Opcional."),
    ("cuidado_temperatura", "Opcional."),
    ("diametro_maceta", "Opcional."),
    ("altura_con_maceta", "Opcional."),
    ("tipo_planta", "Opcional."),
    ("crecimiento", "Lento, Medio o Rápido."),
]


# --------------------------------------------------------------------------- #
# Parsing helpers
# --------------------------------------------------------------------------- #
def _norm_header(value) -> str:
    return slugify(str(value or "")).replace("-", "_")


def _clean_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _money_to_cents(value) -> int | None:
    """Acepta 6800, "6800", "$6.800", "6.800,50" → centavos (int)."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value) * 100))
    text = str(value).strip().replace("$", "").replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return int(round(float(text) * 100))
    except ValueError:
        return None


def _to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(round(float(str(value).replace(",", ".").strip())))
    except ValueError:
        return None


def _to_bool(value) -> bool:
    return str(value or "").strip().lower() in {"si", "sí", "true", "1", "x", "yes", "y"}


def _split_tags(value) -> list[str]:
    if not value:
        return []
    return [t.strip() for t in str(value).replace(";", ",").split(",") if t.strip()]


def _parse_workbook(content: bytes) -> list[dict]:
    """Devuelve una lista de filas {clave_interna: valor_crudo, "_row": nro_excel}."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["Productos"] if "Productos" in wb.sheetnames else wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    col_keys: list[str | None] = []
    for cell in header:
        norm = _norm_header(cell)
        col_keys.append(COLUMN_MAP.get(norm))

    parsed: list[dict] = []
    for idx, raw in enumerate(rows_iter, start=2):
        record: dict = {"_row": idx}
        has_value = False
        for key, cell in zip(col_keys, raw):
            if key is None:
                continue
            if cell is not None and str(cell).strip() != "":
                has_value = True
            record[key] = cell
        if has_value:
            parsed.append(record)
    wb.close()
    return parsed


def _build_product_fields(record: dict, category_id: str | None) -> dict:
    """Campos comunes para crear o actualizar un producto (medida mediana)."""
    currency = (_clean_str(record.get("currency")) or "ARS").upper()
    if currency not in {"ARS", "USD"}:
        currency = "ARS"

    tax = (_clean_str(record.get("tax")) or "iva-21").lower()
    if tax not in {"iva-21", "iva-10", "exento"}:
        tax = "iva-21"

    status = "active" if str(record.get("status") or "").strip().lower() in {
        "activo", "active", "publicado"
    } else "draft"

    weight_kg = record.get("weight_kg")
    weight_grams = None
    if weight_kg not in (None, ""):
        try:
            weight_grams = int(round(float(str(weight_kg).replace(",", ".")) * 1000))
        except ValueError:
            weight_grams = None

    care = {
        "light": _clean_str(record.get("care_light")),
        "water": _clean_str(record.get("care_water")),
        "environment": _clean_str(record.get("care_environment")),
        "temperature": _clean_str(record.get("care_temperature")),
    }
    care = {k: v for k, v in care.items() if v}

    attributes = {
        "pot_diameter": _clean_str(record.get("attr_pot_diameter")),
        "height_with_pot": _clean_str(record.get("attr_height_with_pot")),
        "plant_type": _clean_str(record.get("attr_plant_type")),
        "growth": _clean_str(record.get("attr_growth")),
    }
    attributes = {k: v for k, v in attributes.items() if v}

    return {
        "title": _clean_str(record.get("title")),
        "short_description": _clean_str(record.get("short_description")),
        "description": _clean_str(record.get("description")),
        "price": _money_to_cents(record.get("price")),
        "compare_at_price": _money_to_cents(record.get("compare_at_price")),
        "cost_price": _money_to_cents(record.get("cost_price")),
        "currency": currency,
        "tax": tax,
        "category_id": category_id,
        "stock": _to_int(record.get("stock")) or 0,
        "sku": _clean_str(record.get("sku")),
        "status": status,
        "is_featured": _to_bool(record.get("is_featured")),
        "weight_grams": weight_grams,
        "height_cm": _to_int(record.get("height_cm")),
        "tags": _split_tags(record.get("tags")),
        "care": care,
        "attributes": attributes,
    }


# --------------------------------------------------------------------------- #
# Background job
# --------------------------------------------------------------------------- #
async def _process_import(job_id: str, tenant_id: str, records: list[dict]) -> None:
    db = get_db()
    oid = ObjectId(job_id)

    try:
        cat_docs = await db.categories.find({}).to_list(None)
    except Exception:
        cat_docs = []
    cat_by_slug: dict[str, str] = {}
    for c in cat_docs:
        cat_by_slug[c.get("slug", "")] = str(c["_id"])
        cat_by_slug[slugify(c.get("name", ""))] = str(c["_id"])

    total = len(records)
    processed = created = updated = warnings = errors = 0
    report: list[dict] = []

    async def _flush() -> None:
        await db.import_jobs.update_one(
            {"_id": oid},
            {"$set": {
                "processed": processed, "created": created, "updated": updated,
                "warnings": warnings, "errors": errors, "rows": report,
            }},
        )

    for record in records:
        excel_row = record.get("_row", 0)
        name = _clean_str(record.get("title"))
        raw_cat = _clean_str(record.get("category"))

        category_id: str | None = None
        pending_warning: str | None = None
        if raw_cat:
            category_id = cat_by_slug.get(slugify(raw_cat))
            if category_id is None:
                pending_warning = f"Categoría «{raw_cat}» no encontrada; se importó sin categoría"

        fields = _build_product_fields(record, category_id)

        if not name:
            errors += 1
            processed += 1
            report.append({"row": excel_row, "name": None, "action": "error",
                           "message": "Falta el nombre del producto"})
            await _flush()
            continue
        if not fields["price"] or fields["price"] <= 0:
            errors += 1
            processed += 1
            report.append({"row": excel_row, "name": name, "action": "error",
                           "message": "Falta el precio o es inválido"})
            await _flush()
            continue

        now = datetime.now(UTC)
        sku = fields["sku"]
        existing = None
        if sku:
            existing = await db.products.find_one(
                {"tenant_id": tenant_id, "sku": sku, "deleted_at": None}
            )
        if existing is None:
            existing = await db.products.find_one({
                "tenant_id": tenant_id,
                "deleted_at": None,
                "title": {"$regex": f"^{_regex_escape(name)}$", "$options": "i"},
            })

        try:
            if existing:
                await db.products.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {**fields, "updated_at": now}},
                )
                updated += 1
                action = "updated"
            else:
                doc = {
                    "tenant_id": tenant_id,
                    "tenant_name": tenant_id,
                    **fields,
                    "images": [],
                    "variants": [],
                    "recommended_pot_ids": [],
                    "publish_at": now if fields["status"] == "active" else None,
                    "rating_avg": None,
                    "rating_count": 0,
                    "sold_count": 0,
                    "deleted_at": None,
                    "created_at": now,
                    "updated_at": now,
                }
                await db.products.insert_one(doc)
                created += 1
                action = "created"
        except Exception as exc:  # noqa: BLE001
            errors += 1
            processed += 1
            report.append({"row": excel_row, "name": name, "action": "error",
                           "message": f"No se pudo guardar: {exc}"})
            await _flush()
            continue

        if pending_warning:
            warnings += 1
            report.append({"row": excel_row, "name": name, "action": "warning",
                           "message": pending_warning})
        else:
            report.append({"row": excel_row, "name": name, "action": action, "message": None})

        processed += 1
        if processed % 5 == 0:
            await _flush()

    await db.import_jobs.update_one(
        {"_id": oid},
        {"$set": {
            "status": "completed", "processed": processed, "total": total,
            "created": created, "updated": updated, "warnings": warnings,
            "errors": errors, "rows": report, "finished_at": datetime.now(UTC),
        }},
    )


def _regex_escape(text: str) -> str:
    import re
    return re.escape(text)


# --------------------------------------------------------------------------- #
# Routes
# --------------------------------------------------------------------------- #
@router.get("/template")
async def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append([TEMPLATE_EXAMPLE.get(h, "") for h in TEMPLATE_HEADERS])

    header_col = {h: get_column_letter(i) for i, h in enumerate(TEMPLATE_HEADERS, start=1)}
    for header, letter in header_col.items():
        ws.column_dimensions[letter].width = max(16, len(header) + 4)

    # Desplegables en la celda para los campos con opciones fijas.
    last = TEMPLATE_VALIDATION_ROWS
    for header, choices in DROPDOWN_CHOICES.items():
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(choices) + '"',
            allow_blank=True,
            showErrorMessage=False,
        )
        dv.add(f"{header_col[header]}2:{header_col[header]}{last}")
        ws.add_data_validation(dv)

    info = wb.create_sheet("Instrucciones")
    info.append(["Columna", "Descripción"])
    for cell in info[1]:
        cell.font = Font(bold=True)
    for col, desc in INSTRUCTIONS:
        info.append([col, desc])
    info.append([])
    info.append(["Notas"])
    info.append(["", "Las fotos y las macetas recomendadas no se importan."])
    info.append(["", "Todos los valores corresponden a la medida mediana del producto."])
    info.append(["", "Si el producto ya existe (mismo SKU o mismo nombre) se actualiza."])
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 90

    db = get_db()
    try:
        cats = await db.categories.find({}).sort("name", 1).to_list(None)
    except Exception:
        cats = []
    cat_names = [c["name"] for c in cats if c.get("name")]

    # Hoja auxiliar oculta que alimenta los desplegables con opciones largas
    # (categorías y cuidados, cuyos textos tienen comas).
    listas = wb.create_sheet("Listas")
    listas.sheet_state = "hidden"

    def _add_list_column(col_idx: int, header: str, values: list[str]) -> str:
        letter = get_column_letter(col_idx)
        for row_idx, value in enumerate(values, start=1):
            listas.cell(row=row_idx, column=col_idx, value=value)
        dv = DataValidation(
            type="list",
            formula1=f"Listas!${letter}$1:${letter}${max(len(values), 1)}",
            allow_blank=True,
            showErrorMessage=False,
        )
        dv.add(f"{header_col[header]}2:{header_col[header]}{last}")
        ws.add_data_validation(dv)
        return letter

    col = 1
    if cat_names:
        _add_list_column(col, "categoria", cat_names)
        col += 1
    for care_header, care_values in CARE_CHOICES.items():
        _add_list_column(col, care_header, care_values)
        col += 1

    if cat_names:
        info.append([])
        info.append(["Categorías disponibles"])
        for name in cat_names:
            info.append(["", name])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": 'attachment; filename="plantilla-productos.xlsx"'},
    )


@router.post("", response_model=ImportStartResponse, status_code=202)
async def start_import(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
):
    user_id = require_user_id(request)
    tenant_id = getattr(request.state, "tenant_id", None) or "default"

    filename = file.filename or "productos.xlsx"
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "El archivo debe ser una planilla .xlsx")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "El archivo no puede superar 10MB")

    try:
        records = _parse_workbook(content)
    except Exception:  # noqa: BLE001
        raise HTTPException(
            400, "No se pudo leer la planilla. Usá la plantilla oficial."
        ) from None

    if not records:
        raise HTTPException(400, "La planilla no tiene filas con datos")

    now = datetime.now(UTC)
    db = get_db()
    result = await db.import_jobs.insert_one({
        "tenant_id": tenant_id,
        "created_by": user_id,
        "filename": filename,
        "status": "processing",
        "total": len(records),
        "processed": 0,
        "created": 0,
        "updated": 0,
        "warnings": 0,
        "errors": 0,
        "rows": [],
        "error": None,
        "created_at": now,
        "finished_at": None,
    })
    job_id = str(result.inserted_id)
    background_tasks.add_task(_process_import, job_id, tenant_id, records)
    return {"job_id": job_id}


@router.get("/{job_id}", response_model=ImportJobOut)
async def get_import_job(job_id: str, request: Request):
    db = get_db()
    try:
        oid = ObjectId(job_id)
    except Exception:
        raise HTTPException(400, "ID de importación inválido") from None

    doc = await db.import_jobs.find_one({"_id": oid})
    if not doc:
        raise HTTPException(404, "Importación no encontrada")

    tenant_id = getattr(request.state, "tenant_id", None)
    if tenant_id and doc.get("tenant_id") != tenant_id:
        raise HTTPException(404, "Importación no encontrada")

    return {
        "job_id": str(doc["_id"]),
        "status": doc.get("status", "processing"),
        "filename": doc.get("filename"),
        "total": doc.get("total", 0),
        "processed": doc.get("processed", 0),
        "created": doc.get("created", 0),
        "updated": doc.get("updated", 0),
        "warnings": doc.get("warnings", 0),
        "errors": doc.get("errors", 0),
        "rows": doc.get("rows", []),
        "error": doc.get("error"),
        "created_at": doc.get("created_at"),
        "finished_at": doc.get("finished_at"),
    }
